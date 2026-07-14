#!/usr/bin/env python3
"""Hastane yemek sayfasındaki en güncel XLSX'i indir, parse et, menu.json üret."""
import io, json, re
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

PAGE_URL = 'https://ispartasehir.saglik.gov.tr/TR-471595/yemek-listesi.html'
HEADERS = {'User-Agent': 'Mozilla/5.0 (compatible; hospital-menu-bot/1.0)'}
TURKISH_MONTHS = [
    'ocak','subat','mart','nisan','mayis','haziran',
    'temmuz','agustos','eylul','ekim','kasim','aralik',
]

# ── 1. Find the latest XLSX URL ──────────────────────────────
print('[1] Fetching page...')
html = requests.get(PAGE_URL, timeout=30, headers=HEADERS).text
urls = re.findall(r'https?://[^\s"\'<>]+\.xlsx[^\s"\'<>]*', html)
print(f'    Found {len(urls)} XLSX link(s)')

current_month = TURKISH_MONTHS[datetime.now().month - 1]
best_url = None
for u in reversed(urls):
    if current_month in u.lower():
        best_url = u
        break
best_url = best_url or (urls[-1] if urls else None)
if not best_url:
    raise SystemExit('XLSX linki bulunamadı')
print(f'    Selected: {best_url}')

# ── 2. Download XLSX ─────────────────────────────────────────
print('[2] Downloading XLSX...')
resp = requests.get(best_url, timeout=60, headers=HEADERS)
resp.raise_for_status()
print(f'    {len(resp.content)} bytes')

# ── 3. Save raw XLSX (backup) ────────────────────────────────
xlsx_path = Path('data/yemek_listesi.xlsx')
xlsx_path.parent.mkdir(parents=True, exist_ok=True)
xlsx_path.write_bytes(resp.content)

# ── 4. Parse XLSX → JSON ─────────────────────────────────────
print('[3] Parsing...')
wb = load_workbook(io.BytesIO(resp.content), data_only=True)
ws = wb[wb.sheetnames[0]]

date_rows = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if isinstance(v, datetime):
        date_rows.append((r, v))
print(f'    {len(date_rows)} date anchors found')

menu = {}
for i, (row_idx, date) in enumerate(date_rows):
    end_row = date_rows[i + 1][0] if i + 1 < len(date_rows) else ws.max_row + 1
    date_key = date.strftime('%Y-%m-%d')

    breakfast, lunch, dinner = [], [], []

    for r in range(row_idx, end_row):
        is_date_row = (r == row_idx)

        def add_safe(lst, col):
            v = ws.cell(r, col).value
            if v:
                t = ' '.join(str(v).strip().split())
                if t and t not in lst:
                    lst.append(t)

        if not is_date_row:
            add_safe(breakfast, 1)   # soup / extra in column A
        add_safe(breakfast, 8)       # column H = breakfast items
        add_safe(lunch, 16)          # column P = lunch items
        add_safe(dinner, 24)         # column X = dinner items

    menu[date_key] = {
        'breakfast': breakfast,
        'lunch': lunch,
        'dinner': dinner,
    }

# ── 5. Save JSON ─────────────────────────────────────────────
json_path = Path('data/menu.json')
json_path.parent.mkdir(parents=True, exist_ok=True)
json_path.write_text(json.dumps(menu, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'    Saved {len(menu)} days → {json_path}')

# ── 6. Print today's menu ────────────────────────────────────
today = datetime.now().strftime('%Y-%m-%d')
if today in menu:
    print(f'\n[✓] Today ({today}):')
    for meal, label in [('breakfast', 'Sabah'), ('lunch', 'Öğle'), ('dinner', 'Akşam')]:
        items = ' • '.join(menu[today][meal])
        print(f'    {label}: {items}')
else:
    print(f'[!] {today} not found')